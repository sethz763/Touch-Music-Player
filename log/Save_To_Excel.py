import datetime
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import  Border, Side, Alignment, Font
from fleep import *
import fleep
from persistence.SaveSettings import SaveSettings
from PySide6.QtCore import QObject, Signal, QTimer
from PySide6.QtWidgets import QPushButton, QVBoxLayout, QGridLayout, QWidget, QApplication, QHBoxLayout, QSpacerItem, QRadioButton, QSlider, QLabel, QComboBox, QMainWindow, QLineEdit, QSpinBox, QFileDialog, QMessageBox, QDialog
from PySide6 import QtCore

import shutil
import os
from inspect import currentframe, getframeinfo

class Save_To_Excel(QObject):
    
    # Signal emitted when a new log entry is added
    log_entry_added = Signal(dict)  # emits the log_data dict
    
    def __init__(self, filename:str="/log.xlsx", title:str="Music Log", parent=None):
        super().__init__(parent)  # Initialize QObject parent class
        
        try:
            self.parent = parent
            self.app_dir = os.path.abspath(os.path.dirname(__file__))
            self.save_settings = SaveSettings('log_settings.json')
            self.excel_file_settings = self.save_settings.get_settings()

            # Initialize defaults
            self.filename = ""  # Will be set to passed filename or from settings
            self.title = title
            self.merged_cells = []
            self.file_found = False
            self.music_log_excel_file = None
            self.music_log_sheet = None
            
            self.thin = Side(border_style="thin", color="000000")
            self.thin_borders = Border(top=self.thin, bottom=self.thin, left=self.thin, right=self.thin)

            # Try to load from saved settings first
            if 'filename' in self.excel_file_settings and self.excel_file_settings['filename']:
                saved_filename = self.excel_file_settings['filename']
                if self.isExcelFile(saved_filename):
                    self.filename = saved_filename
                    self.music_log_excel_file = load_workbook(filename=self.filename)
                    self.music_log_sheet = self.music_log_excel_file.active
                    self.merged_cells = list(self.music_log_sheet.merged_cells)
                    self.file_found = True

                    if 'show_name' in self.excel_file_settings:
                        self.title = self.excel_file_settings['show_name']
                    # Successfully loaded, skip to end
                else:
                    # Saved file doesn't exist or isn't valid Excel, create new
                    self.start_new_log(filename=filename, title=title)
            else:
                # No saved settings, create new log with provided filename
                self.start_new_log(filename=filename, title=title)

            #get system date and time
            self.date = datetime.datetime.now()
            self.time = datetime.datetime.now()
            self.date = self.date.strftime('%m-%d-%Y')
            self.time = self.time.strftime('%H:%M:%S')
            
            # self.auto_save_timer = QTimer()
            # self.auto_save_duration = 10000 #10 seconds
            # self.auto_save_timer.setInterval(self.auto_save_duration)
            # self.auto_save_timer.timeout.connect(self.save)
            # self.auto_save_timer.start()

            # Perf instrumentation
            self._perf_threshold_ms = 5.0
            # Save debounce/coalescing: avoid doing a full workbook save per cue finish.
            self._save_debounce_ms = 250
            self._save_timer = QTimer(self)
            self._save_timer.setSingleShot(True)
            self._save_timer.timeout.connect(self._run_debounced_save)

            self._save_in_progress = False
            self._save_dirty = False
            self._save_requested_at = None  # perf_counter timestamp
            self._save_request_count = 0

            # Backup I/O can be expensive; rate-limit backups during bursts.
            self._last_backup_ts = 0.0
            
        except Exception as e:
            info = getframeinfo(currentframe())
            print(f'{e}:{info.filename}:{info.lineno}')  
       
    def update_log(self, log_data={}):
        start = time.perf_counter()
        try:
            # Warn if no filename is set
            if not self.filename or self.filename == '':
                print("Warning: Cannot log entry - no Excel file selected. Please create or load a log file first.")
                return
            
            self.date = datetime.datetime.now()
            self.date = self.date.strftime('%m-%d-%Y')
            self.log_data = log_data
            date = self.date
            artist = self.log_data['ARTIST']
            song = self.log_data['SONG']
            filename = self.log_data['FILENAME']
            time_start = self.log_data['TIME_START']
            time_end = self.log_data['TIME_END']
            duration_played = self.log_data['DURATION_PLAYED']

            data = [artist, song, filename, time_start, time_end, duration_played]
            self.music_log_sheet.append(data)

            current_row = self.music_log_sheet.max_row
            for cell in self.music_log_sheet.iter_cols(min_row=current_row, max_row=current_row, min_col=0, max_col=6, values_only=False):
                cell[0].border = self.thin_borders
                cell[0].font = Font(name='Arial', size=10, bold=False)

            # Debounced save to avoid blocking GUI repeatedly during bursts
            self._queue_save_async()
            
            # Emit signal so listeners (e.g., logging dialog) can refresh
            self.log_entry_added.emit(log_data)

            elapsed_ms = (time.perf_counter() - start) * 1000
            if elapsed_ms > self._perf_threshold_ms:
                try:
                    rows = self.music_log_sheet.max_row if self.music_log_sheet is not None else -1
                except Exception:
                    rows = -1
                from log.perf import perf_print

                perf_print(f"[PERF] Save_To_Excel.update_log: {elapsed_ms:.2f}ms rows={rows} save_requests_pending={self._save_request_count}")

        except Exception as e:
            info = getframeinfo(currentframe())
            print(f'{e}:{info.filename}:{info.lineno}')
    
    def _queue_save_async(self):
        """Debounce saves via a single QTimer to coalesce bursts into one save."""
        try:
            self._save_request_count += 1
            if self._save_requested_at is None:
                self._save_requested_at = time.perf_counter()
            self._save_dirty = True
        except Exception:
            pass

        # If a save is currently running, just mark dirty; completion will reschedule.
        if getattr(self, "_save_in_progress", False):
            return

        # Restart debounce timer: last update wins.
        try:
            self._save_timer.start(self._save_debounce_ms)
        except Exception:
            # Fallback: if timer can't start, do an immediate save.
            self.save()

    def _run_debounced_save(self):
        if getattr(self, "_save_in_progress", False):
            self._save_dirty = True
            return

        self._save_in_progress = True

        requested_at = self._save_requested_at
        req_count = self._save_request_count
        # Clear request tracking for this save run. New requests during save will re-mark dirty.
        self._save_requested_at = None
        self._save_request_count = 0
        self._save_dirty = False

        delay_ms = 0.0
        if requested_at is not None:
            delay_ms = (time.perf_counter() - requested_at) * 1000

        t0 = time.perf_counter()
        try:
            self.save()
        finally:
            save_ms = (time.perf_counter() - t0) * 1000
            self._save_in_progress = False
            if save_ms > 10.0 or delay_ms > 10.0 or req_count > 1:
                print(
                    f"[PERF] Save_To_Excel.save(debounced): save={save_ms:.2f}ms delay={delay_ms:.2f}ms coalesced={req_count}"
                )

            # If more updates came in while saving, schedule another debounced save.
            if self._save_dirty or self._save_request_count > 0:
                if self._save_requested_at is None:
                    self._save_requested_at = time.perf_counter()
                try:
                    self._save_timer.start(self._save_debounce_ms)
                except Exception:
                    pass

    def set_filename(self, filename):
        self.filename = filename
        self.save_settings.set_setting('filename', self.filename)
        self.save_settings.save_settings()

    def set_show_name(self, show_name):
        self.title = show_name
        self.save_settings.set_setting('show_name', self.title)
        self.save_settings.save_settings()

    def start_new_log(self, filename="log.xlsx", title="Music Log"):
        try:
            self.music_log_excel_file = Workbook()
            self.music_log_sheet = self.music_log_excel_file.active
            #get system date and time
            self.date = datetime.datetime.now()
            self.time = datetime.datetime.now()
            self.date = self.date.strftime('%m-%d-%Y')
            self.time = self.time.strftime('%H:%M:%S')

            self.filename = filename
            self.title = title
            self.music_log_sheet['A1'] = f'{self.title} | DATE: {self.date}'
            self.music_log_sheet.merge_cells('A1:F1')
            self.merged_cells = list(self.music_log_sheet.merged_cells)
            self.music_log_sheet.column_dimensions
            cell = self.music_log_sheet['A1']
            cell.alignment = Alignment(horizontal='center')
            thick = Side(border_style="thick", color="000000")
            borders = Border(top=thick, bottom=thick, left=thick, right=thick)
            cell.border = borders
            cell.font = Font(name='Arial', size=12, bold=True)

            #add heading columns
            self.headers = ['ARTIST','SONG','FILENAME', 'TIME_START', 'TIME_END', 'DURATION_PLAYED']
            self.music_log_sheet.append(self.headers)
            #make headers bold and outlined
            cell = self.music_log_sheet['A2']
            self.music_log_sheet.column_dimensions['A'].width = 25
            self.music_log_sheet.column_dimensions['B'].width = 25
            self.music_log_sheet.column_dimensions['C'].width = 25
            self.music_log_sheet.column_dimensions['D'].width = 15
            self.music_log_sheet.column_dimensions['E'].width = 15
            self.music_log_sheet.column_dimensions['F'].width = 18
            
            for cell in self.music_log_sheet.iter_cols(min_row=2, max_row=2, min_col=0, max_col=6, values_only=False):
                cell[0].border = self.thin_borders
                cell[0].font = Font(name='Arial', size=10, bold=True)

            self.save()
            self.set_filename(self.filename)

        except Exception as e:
            print(e)
            
    def load(self, filename):
        if not filename or filename == '':
            print("Warning: Cannot load log file - filename is empty")
            return
        
        if self.isExcelFile(filename):
            self.filename = filename
            self.music_log_excel_file = load_workbook(filename=self.filename)
            self.music_log_sheet = self.music_log_excel_file.active
            self.file_found = True
            self.title = self.music_log_sheet['A1'].value
            self.merged_cells = [self.music_log_sheet.merged_cells]
            self.set_filename(self.filename)
            # Only emit signals if parent is set
            if self.parent and hasattr(self.parent, 'get_info_signal'):
                self.parent.get_info_signal.emit(self.filename, self.title, self.get_num_entries())
                self.parent.get_sheet_signal.emit(list(self.music_log_sheet.values), self.merged_cells)
        else:
            print(f"Warning: File is not a valid Excel file: {filename}")

    def save(self):
        try:
            # Don't save if no filename is set
            if not self.filename or self.filename == '':
                print(f"Warning: Cannot save log file - no filename set")
                return
            
            # save_backup first (rate-limited during bursts)
            backup_dir = "backup_logs"
            if not os.path.exists(backup_dir):
                os.mkdir(backup_dir)

            now = time.perf_counter()
            do_backup = (now - getattr(self, "_last_backup_ts", 0.0)) > 2.0
            if do_backup and os.path.exists(self.filename):
                backup_path = backup_dir + "/" + os.path.basename(self.filename)
                shutil.copy(self.filename, backup_path)
                self._last_backup_ts = now
            
            self.music_log_excel_file.save(self.filename)
            # self.music_log_excel_file = load_workbook(filename=self.filename)
            # self.music_log_sheet = self.music_log_excel_file.active            
            
        except Exception as e:
            info = getframeinfo(currentframe())
            print(f'{e}:{info.filename}:{info.lineno}')  

    def isExcelFile(self, file):
        # Check if file path is empty or None
        if not file or file == '':
            return False
        
        try:
            with open(file, 'rb') as check_file:
                    info = fleep.get(check_file.read(128))
                    if info.type == ['document', 'archive', 'executable']:
                        return True
                    else:
                        return False
        except Exception as e:
            info = getframeinfo(currentframe())
            print(f'{e}:{info.filename}:{info.lineno}')
            return False  

    def get_num_entries(self):
        first_data_row = 2
        num_entries = self.music_log_sheet.max_row - first_data_row
        return num_entries

    def clear(self):
        try:
            first_data_row = 3
            rows_to_delete = self.music_log_sheet.max_row - first_data_row
            self.music_log_sheet.delete_rows(first_data_row, rows_to_delete+1) 
            self.save()
            
        except:
            pass