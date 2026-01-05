from __future__ import annotations

import csv
import os
import queue
import time
from dataclasses import dataclass, asdict
from typing import Any, Dict, Iterable, List, Optional, Tuple

import multiprocessing as mp

from PySide6.QtCore import QObject, Signal


RAW_FIELDS: Tuple[str, ...] = (
    "ARTIST",
    "SONG",
    "FILENAME",
    "TIME_START",
    "TIME_END",
    "DURATION_PLAYED",
    "CUE_ID",
    "TRACK_ID",
    "IN_FRAME",
    "OUT_FRAME",
    "GAIN_DB",
    "DURATION_SECONDS",
    "FADE_IN_MS",
    "FADE_OUT_MS",
    "END_REASON",
)


@dataclass(frozen=True)
class LogWriterConfig:
    csv_path: str
    xlsx_path: str
    title: str = "Cue Log"
    save_debounce_ms: int = 750
    backup_interval_s: float = 2.0


# ---------- Worker process ----------

def _ensure_parent_dir(path: str) -> None:
    parent = os.path.dirname(os.path.abspath(path))
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)


def _write_csv_header_if_needed(csv_path: str, fieldnames: Iterable[str]) -> None:
    _ensure_parent_dir(csv_path)
    needs_header = (not os.path.exists(csv_path)) or (os.path.getsize(csv_path) == 0)
    if not needs_header:
        return
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(fieldnames), extrasaction="ignore")
        writer.writeheader()


def _append_csv_row(csv_path: str, row: Dict[str, Any]) -> None:
    _write_csv_header_if_needed(csv_path, RAW_FIELDS)
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(RAW_FIELDS), extrasaction="ignore")
        writer.writerow({k: row.get(k, "") for k in RAW_FIELDS})


def _open_or_create_workbook(xlsx_path: str, title: str):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    _ensure_parent_dir(xlsx_path)

    if os.path.exists(xlsx_path):
        wb = load_workbook(filename=xlsx_path)
    else:
        wb = Workbook()

    # Ensure sheets
    if "Cue Log" in wb.sheetnames:
        cue_ws = wb["Cue Log"]
    else:
        cue_ws = wb.active
        cue_ws.title = "Cue Log"

    raw_ws = wb["Raw"] if "Raw" in wb.sheetnames else wb.create_sheet("Raw")

    # Prepare header/title rows if empty
    thin = Side(border_style="thin", color="000000")
    thin_borders = Border(top=thin, bottom=thin, left=thin, right=thin)

    if cue_ws.max_row < 2:
        cue_ws["A1"] = f"{title}"
        cue_ws.merge_cells("A1:F1")
        cue_ws["A1"].alignment = Alignment(horizontal="center")
        cue_ws["A1"].border = thin_borders
        cue_ws["A1"].font = Font(name="Arial", size=12, bold=True)
        cue_ws.append(["ARTIST", "SONG", "FILENAME", "TIME_START", "TIME_END", "DURATION_PLAYED"])

    if raw_ws.max_row < 1:
        raw_ws.append(list(RAW_FIELDS))

    return wb, cue_ws, raw_ws


def _append_excel_rows(cue_ws, raw_ws, row: Dict[str, Any]) -> None:
    # Human-friendly sheet (first 6 columns)
    cue_ws.append(
        [
            row.get("ARTIST", ""),
            row.get("SONG", ""),
            row.get("FILENAME", ""),
            row.get("TIME_START", ""),
            row.get("TIME_END", ""),
            row.get("DURATION_PLAYED", ""),
        ]
    )

    # Raw sheet (full)
    raw_ws.append([row.get(k, "") for k in RAW_FIELDS])


def log_writer_main(cmd_q: "mp.Queue", cfg_dict: Dict[str, Any]) -> None:
    """Process entry point. Owns CSV and XLSX writes; GUI never touches openpyxl saves."""
    cfg = LogWriterConfig(**cfg_dict)

    wb = None
    cue_ws = None
    raw_ws = None

    dirty = False
    last_change = 0.0
    last_save = 0.0
    last_backup = 0.0

    def _ensure_open():
        nonlocal wb, cue_ws, raw_ws
        if wb is None:
            wb, cue_ws, raw_ws = _open_or_create_workbook(cfg.xlsx_path, cfg.title)

    def _backup_if_needed() -> None:
        nonlocal last_backup
        now = time.perf_counter()
        if (now - last_backup) < cfg.backup_interval_s:
            return
        last_backup = now
        try:
            if not os.path.exists(cfg.xlsx_path):
                return
            backup_dir = os.path.join(os.path.dirname(os.path.abspath(cfg.xlsx_path)) or ".", "backup_logs")
            os.makedirs(backup_dir, exist_ok=True)
            backup_path = os.path.join(backup_dir, os.path.basename(cfg.xlsx_path))
            # copy via read/write to avoid shutil import overhead here
            with open(cfg.xlsx_path, "rb") as src, open(backup_path, "wb") as dst:
                dst.write(src.read())
        except Exception:
            pass

    def _save(force: bool = False) -> None:
        nonlocal dirty, last_save
        if wb is None:
            return
        if not dirty and not force:
            return
        _backup_if_needed()
        wb.save(cfg.xlsx_path)
        dirty = False
        last_save = time.perf_counter()

    def _clear_files() -> None:
        nonlocal wb, cue_ws, raw_ws, dirty
        # Reset CSV
        _ensure_parent_dir(cfg.csv_path)
        with open(cfg.csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(RAW_FIELDS), extrasaction="ignore")
            writer.writeheader()

        # Reset workbook
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        cue_ws_local = wb.create_sheet("Cue Log")
        raw_ws_local = wb.create_sheet("Raw")

        cue_ws_local["A1"] = f"{cfg.title}"
        cue_ws_local.merge_cells("A1:F1")
        cue_ws_local.append(["ARTIST", "SONG", "FILENAME", "TIME_START", "TIME_END", "DURATION_PLAYED"])
        raw_ws_local.append(list(RAW_FIELDS))

        cue_ws = cue_ws_local
        raw_ws = raw_ws_local
        dirty = True
        _save(force=True)

    debounce_s = max(0.0, cfg.save_debounce_ms / 1000.0)

    while True:
        # If dirty and enough time passed since last change, save even if events keep trickling.
        now = time.perf_counter()
        if dirty and (now - last_change) >= debounce_s:
            try:
                _save()
            except Exception:
                pass

        try:
            msg = cmd_q.get(timeout=0.05)
        except queue.Empty:
            continue

        if not msg:
            continue

        mtype = msg[0]

        if mtype == "shutdown":
            try:
                _save(force=True)
            except Exception:
                pass
            return

        if mtype == "flush":
            try:
                _save(force=True)
            except Exception:
                pass
            continue

        if mtype == "set_paths":
            payload = msg[1] or {}
            cfg = LogWriterConfig(
                csv_path=payload.get("csv_path", cfg.csv_path),
                xlsx_path=payload.get("xlsx_path", cfg.xlsx_path),
                title=payload.get("title", cfg.title),
                save_debounce_ms=payload.get("save_debounce_ms", cfg.save_debounce_ms),
                backup_interval_s=payload.get("backup_interval_s", cfg.backup_interval_s),
            )
            # Re-open workbook lazily on next use.
            wb = None
            cue_ws = None
            raw_ws = None
            dirty = False
            continue

        if mtype == "clear":
            try:
                _clear_files()
            except Exception:
                pass
            continue

        if mtype == "append":
            row: Dict[str, Any] = msg[1] or {}
            try:
                _append_csv_row(cfg.csv_path, row)
            except Exception:
                pass
            try:
                _ensure_open()
                _append_excel_rows(cue_ws, raw_ws, row)
                dirty = True
                last_change = time.perf_counter()
            except Exception:
                pass
            continue


# ---------- GUI-side proxy ----------

class _SheetSnapshot:
    def __init__(self, values: List[Tuple[Any, ...]]):
        self.values = values
        self.merged_cells: List[Any] = []


class AsyncCsvExcelLogger(QObject):
    """GUI-safe logger: enqueue rows to a worker process, but keep a local snapshot for the log dialog."""

    log_entry_added = Signal(dict)

    def __init__(
        self,
        csv_path: str = "cue_log.csv",
        xlsx_path: str = "cue_log.xlsx",
        title: str = "Cue Log",
        save_debounce_ms: int = 750,
        preload_snapshot: bool = True,
        enabled: bool = True,
        parent: Optional[QObject] = None,
    ) -> None:
        super().__init__(parent)

        self._ctx = mp.get_context("spawn")
        self._q: "mp.Queue" = self._ctx.Queue()
        self._proc = self._ctx.Process(
            target=log_writer_main,
            args=(self._q, asdict(LogWriterConfig(csv_path=csv_path, xlsx_path=xlsx_path, title=title, save_debounce_ms=save_debounce_ms))),
            daemon=True,
        )
        self._proc.start()

        # Minimal compatibility for Log_Settings_Window.refresh()
        self.music_log_sheet = _SheetSnapshot(
            values=[
                (title, None, None, None, None, None),
                ("ARTIST", "SONG", "FILENAME", "TIME_START", "TIME_END", "DURATION_PLAYED"),
            ]
        )

        self._csv_path = csv_path
        self._xlsx_path = xlsx_path
        self._title = title
        self._enabled = bool(enabled)

        if preload_snapshot:
            self._preload_snapshot()

    def _reset_snapshot(self) -> None:
        self.music_log_sheet.values = [
            (self._title, None, None, None, None, None),
            ("ARTIST", "SONG", "FILENAME", "TIME_START", "TIME_END", "DURATION_PLAYED"),
        ]

    def _preload_snapshot(self) -> None:
        """Populate the dialog-visible snapshot from existing log files.

        Prefers CSV because it's fast and doesn't require openpyxl.
        Falls back to XLSX (read-only) if CSV isn't present.
        """
        self._reset_snapshot()

        # 1) Fast path: CSV
        try:
            if self._csv_path and os.path.exists(self._csv_path) and os.path.getsize(self._csv_path) > 0:
                with open(self._csv_path, "r", newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        self.music_log_sheet.values.append(
                            (
                                row.get("ARTIST", ""),
                                row.get("SONG", ""),
                                row.get("FILENAME", ""),
                                row.get("TIME_START", ""),
                                row.get("TIME_END", ""),
                                row.get("DURATION_PLAYED", ""),
                            )
                        )
                return
        except Exception:
            # If CSV parsing fails, try XLSX.
            pass

        # 2) Fallback: XLSX (read-only)
        try:
            if self._xlsx_path and os.path.exists(self._xlsx_path) and os.path.getsize(self._xlsx_path) > 0:
                from openpyxl import load_workbook

                wb = load_workbook(filename=self._xlsx_path, read_only=True, data_only=True)
                ws = wb["Cue Log"] if "Cue Log" in wb.sheetnames else wb.active
                rows = list(ws.values)
                if rows:
                    # Keep our title row + headers, but load the data rows after the header.
                    # Expected layout: row0 title, row1 headers.
                    data_rows = rows[2:] if len(rows) > 2 else []
                    for r in data_rows:
                        if not r:
                            continue
                        padded = list(r) + [""] * (6 - len(r))
                        self.music_log_sheet.values.append(tuple(padded[:6]))
        except Exception:
            pass

    def set_filename(self, filename: str) -> None:
        # Keep legacy naming; this is actually the xlsx path.
        self._xlsx_path = filename
        try:
            base, _ext = os.path.splitext(filename)
            self._csv_path = base + ".csv"
        except Exception:
            pass
        self._q.put(("set_paths", {"xlsx_path": filename, "csv_path": self._csv_path}))
        self._preload_snapshot()

    def set_title(self, title: str) -> None:
        """Update the title used for new/cleared logs and the dialog snapshot header.

        Does not clear or rewrite existing log files.
        """
        self._title = str(title or self._title)
        try:
            if self.music_log_sheet.values and len(self.music_log_sheet.values) >= 1:
                # Keep the merged-style row shape expected by Log_Settings_Window.
                self.music_log_sheet.values[0] = (self._title, None, None, None, None, None)
        except Exception:
            pass
        try:
            self._q.put(("set_paths", {"title": self._title}))
        except Exception:
            pass

    def start_new_log(self, filename: str = "cue_log.xlsx", title: str = "Cue Log") -> None:
        self._title = title
        self._xlsx_path = filename
        try:
            base, _ext = os.path.splitext(filename)
            self._csv_path = base + ".csv"
        except Exception:
            pass
        self._q.put(("set_paths", {"xlsx_path": filename, "csv_path": self._csv_path, "title": title}))
        self._q.put(("clear", None))

        self._reset_snapshot()

    def load(self, filename: str) -> None:
        # For now, treat load as switching output path; UI snapshot remains local.
        if filename:
            self.set_filename(filename)

    def save(self) -> None:
        self._q.put(("flush", None))

    def clear_sheet(self) -> None:
        self._q.put(("clear", None))
        self.music_log_sheet.values = [
            (self._title, None, None, None, None, None),
            ("ARTIST", "SONG", "FILENAME", "TIME_START", "TIME_END", "DURATION_PLAYED"),
        ]

    def update_log(self, log_data: Dict[str, Any] = None) -> None:
        if not getattr(self, "_enabled", True):
            return
        if not log_data:
            return

        # Update local snapshot (dialog view) with the human-friendly row.
        self.music_log_sheet.values.append(
            (
                log_data.get("ARTIST", ""),
                log_data.get("SONG", ""),
                log_data.get("FILENAME", ""),
                log_data.get("TIME_START", ""),
                log_data.get("TIME_END", ""),
                log_data.get("DURATION_PLAYED", ""),
            )
        )

        # Enqueue full row for CSV + Raw sheet.
        self._q.put(("append", log_data))

        # Notify UI listeners.
        self.log_entry_added.emit(log_data)

    def set_logging_enabled(self, enabled: bool) -> None:
        """Enable/disable enqueueing new log entries.

        Note: does not delete existing logs; it only stops producing new rows.
        """
        self._enabled = bool(enabled)

    def close(self, timeout_s: float = 2.0) -> None:
        try:
            self._q.put(("shutdown", None))
        except Exception:
            pass
        try:
            self._proc.join(timeout=timeout_s)
        except Exception:
            pass
        if self._proc.is_alive():
            try:
                self._proc.terminate()
            except Exception:
                pass
