import openpyxl.workbook
import openpyxl.worksheet
from fleep import *
import fleep
import openpyxl
from persistence.SaveSettings import SaveSettings
from PySide6.QtCore import QObject, Signal, Slot
from PySide6.QtWidgets import (QPushButton, 
                             QVBoxLayout, 
                             QGridLayout, 
                             QWidget, 
                             QApplication, 
                             QHBoxLayout, 
                             QSpacerItem, 
                             QLabel,
                             QLineEdit,
                             QFileDialog, 
                             QMessageBox, 
                             QDialog, 
                             QTableView)
from PySide6 import QtCore

from inspect import currentframe, getframeinfo

class Log_Settings_Window(QWidget):
    
    create_sheet_signal = Signal(str, str) #filename, showname
    clear_sheet_signal = Signal()
    get_sheet_info_signal = Signal()
    load_sheet_signal = Signal(str) #filename
    enable_disable_logging_signal = Signal(bool)
    get_sheet_signal = Signal()
    
    def __init__(self, width=650, height=360, *args, excel_logger=None, **kwargs):
        super().__init__(*args, **kwargs)
        self.setWindowFlags(QtCore.Qt.WindowType.WindowStaysOnTopHint)
        self.setWindowTitle('Settings')
        self.setFixedSize(width, height)
        self.move(100, 100)
        
        # Store reference to Excel logger so we can get sheet data
        self.excel_logger = excel_logger
        
        self.settings_obj = SaveSettings('log_settings.json')
        self.settings = self.settings_obj.get_settings()
        

        self.save_as_btn = QPushButton('Save As')
        self.save_as_btn.setFixedWidth(60)
        self.load_btn = QPushButton('Load Sheet')
        self.load_btn.setFixedWidth(70)
        self.file_name_label = QLineEdit()
        self.filename = ''
        self.file_name_label.setStyleSheet("background-color: 'white';")
        self.file_name_label.setFixedSize(425,25)
        self.show_name_label = QLabel('Title')
        self.title_line_edit = QLineEdit()
        self.create_sheet_btn = QPushButton('Create Sheet')
        self.create_sheet_btn.setFixedWidth(100)
        self.clear_sheet_btn = QPushButton('Clear Sheet')
        self.clear_sheet_btn.setFixedWidth(100)
        self.clear_sheet_btn.setDisabled(False)
        self.status_label = QLabel('')
        self.status_label.setFixedHeight(20)
        
        self.enable_logging_button = QPushButton()
        self.enable_logging_button.setFixedWidth(120)
        self.enable_logging_button.clicked.connect(self.enable_disable_logging)
        if 'logging_enabled' in self.settings:
            print(f"logging_enabled found, {self.settings['logging_enabled']}")
            self.logging_enabled = self.settings['logging_enabled']
            if self.logging_enabled:
                self.enable_disable_logging_signal.emit(True)
                self.enable_logging_button.setText('LOGGING ENABLED')
            else:
                self.enable_disable_logging_signal.emit(False)
                self.enable_logging_button.setText('LOGGING DISABLED')
        
        else:
            self.logging_enabled = False
            self.enable_logging_button.setText('LOGGING DISABLED')
            self.enable_disable_logging_signal.emit(False)
            self.settings_obj.set_setting('logging_enabled', self.logging_enabled)
            self.settings_obj.save_settings()
            
        
        self.table_view = QTableView()
        self.table_view.setFixedSize(int(width*.96), int(height*.6))
        self.merged_cells = None
        
        class Log_Info(object):
            def __init__(self):
                self.filename = ''
                self.title = '' 
                self.num_entries = 0
            
        self.log_info = Log_Info()

        self.main_layout = QVBoxLayout()
        self.top_layout = QHBoxLayout()
        self.middle_layout = QHBoxLayout()
        self.bottom_layout = QHBoxLayout()
        self.table_view_layout = QVBoxLayout()

        self.main_layout.addLayout(self.top_layout)
        self.main_layout.addLayout(self.middle_layout)
        self.main_layout.addLayout(self.bottom_layout)
        self.main_layout.addLayout(self.table_view_layout)
        self.main_layout.addWidget(self.status_label)
        self.top_layout.addWidget(self.save_as_btn)
        self.top_layout.addWidget(self.load_btn)
        self.top_layout.addWidget(self.file_name_label)
        self.middle_layout.addWidget(self.show_name_label)
        self.middle_layout.addWidget(self.title_line_edit)
        self.bottom_layout.addWidget(self.clear_sheet_btn, alignment=QtCore.Qt.AlignmentFlag.AlignLeft)
        self.bottom_layout.addWidget(self.create_sheet_btn, alignment=QtCore.Qt.AlignmentFlag.AlignRight)
        self.bottom_layout.addWidget(self.enable_logging_button)
        self.table_view_layout.addWidget(self.table_view)
        
        self.save_as_btn.clicked.connect(self.save_as)
        self.load_btn.clicked.connect(self.load)
        self.title_line_edit.editingFinished.connect(self.set_title)
        self.create_sheet_btn.clicked.connect(self.create_sheet)
        self.clear_sheet_btn.clicked.connect(self.clear_sheet)

        self.setLayout(self.main_layout)

    def refresh(self):
        """Refresh the table view with current sheet data from the Excel logger."""
        if self.excel_logger is None:
            return
        
        try:
            # Get current sheet data from Excel logger
            sheet_data = list(self.excel_logger.music_log_sheet.values)
            print(f"DEBUG: sheet_data type: {type(sheet_data)}, length: {len(sheet_data)}")
            print(f"DEBUG: first 3 rows: {sheet_data[:3] if sheet_data else 'empty'}")
            merged_cells = list(self.excel_logger.music_log_sheet.merged_cells) if self.excel_logger.music_log_sheet.merged_cells else []
            
            # Update the table view with the data
            self.receive_sheet_data(sheet_data, merged_cells)
        except Exception as e:
            print(f"Error refreshing sheet data: {e}")
            import traceback
            traceback.print_exc()
            
    @Slot(str, str, int) #filename, title, num entries    
    def receive_sheet_info(self, filename:str, title:str, num_entries:int):
        self.log_info.filename = filename
        self.log_info.title = title
        self.file_name_label.setText(self.log_info.filename)
        self.title_line_edit.setText(self.log_info.title)
        self.status_label.setText(f'Log File Already Exists with {num_entries} entries')
        if num_entries > 0:
            self.clear_sheet_btn.setEnabled(True)
        else:
            self.clear_sheet_btn.setEnabled(False)
            
    @Slot(list, list)
    def receive_sheet_data(self, data:list, merged_cells:list):
        try:
            print(f"DEBUG receive_sheet_data: received {len(data)} rows")
            for i, row in enumerate(data[:5]):  # Print first 5 rows for debugging
                print(f"  Row {i}: {row}")
            
            table_model = TableModel(data, merged_cells)
            self.table_view.setModel(table_model)
            
            # Only try to merge cells if data has rows and first row has content
            if data and len(data) > 0 and len(data[0]) > 5 and data[0][0] and data[0][1:5] == (None, None, None, None):
                self.table_view.setSpan(0, 0, 1, 5)
            
            self.table_view.horizontalHeader().hide()
            self.table_view.verticalHeader().hide()
            
            if len(data) > 1:  # More than just headers
                self.clear_sheet_btn.setEnabled(True)
        except Exception as e:
            info = getframeinfo(currentframe())
            print(f'{e}:{info.filename}:{info.lineno}')
            import traceback
            traceback.print_exc()  

    def save_as(self):
        filter = 'excel files (*.xlsx)'
        self.filename = QFileDialog().getSaveFileName(filter=filter)
        self.file_name_label.setText(self.filename[0])
        self.log_info.filename = self.filename[0]
        self.create_sheet_btn.setEnabled(True)
        
    def load(self):
        filter = 'excel files (*.xlsx)'
        self.filename = QFileDialog().getOpenFileName(filter=filter)
        self.file_name_label.setText(self.filename[0])
        self.log_info.filename = self.filename[0]
        self.load_sheet_signal.emit(self.log_info.filename)
        self.create_sheet_btn.setEnabled(False)
        self.clear_sheet_btn.setEnabled(True)

    def set_title(self):
        self.show_name_label = self.title_line_edit.text()
        self.log_info.title = self.title_line_edit.text()

    def create_sheet(self):
        try:
            filename = self.filename[0]
        
        except:
            self.filename =[]
            self.filename.append('log.xlsx')
            
        self.create_sheet_signal.emit(self.filename[0], self.title_line_edit.text())
        self.status_label.setText('New Log Created')
        self.log_info.title = self.title_line_edit.text()
        self.log_info.filename = self.filename[0]
    
    def clear_sheet(self):
        self.dialog_box = QDialog()
        self.dialog_box.setFixedSize(200, 100)
        self.dialog_box.setWindowTitle('Really Clear Log File?')
        self.dialog_box.setWindowFlags(QtCore.Qt.WindowType.WindowStaysOnTopHint)
        label = QLabel('Are you sure you want to\nclear the log file?')
        layout = QVBoxLayout()
        button_layout = QHBoxLayout()
        ok_btn = QPushButton('OK')
        ok_btn.clicked.connect(self.dialog_box.accept)
        cancel_btn = QPushButton('CANCEL')
        cancel_btn.clicked.connect(self.dialog_box.reject)
        self.dialog_box.setLayout(layout)
        layout.addWidget(label, alignment=QtCore.Qt.AlignmentFlag.AlignHCenter)
        button_layout.addWidget(ok_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        if self.dialog_box.exec() == True:
            self.clear_sheet_signal.emit()
            self.status_label.setText('All Log Entries Cleared')
        
    def enable_disable_logging(self):
        if self.logging_enabled == False:
            self.logging_enabled = True
            self.enable_logging_button.setText('LOGGING ENABLED')
            self.enable_disable_logging_signal.emit(True)
            self.settings_obj.set_setting('logging_enabled', self.logging_enabled)
            self.settings_obj.save_settings()
        else:
            self.logging_enabled = False
            self.enable_logging_button.setText('LOGGING DISABLED')
            self.enable_disable_logging_signal.emit(False)
            self.settings_obj.set_setting('logging_enabled', self.logging_enabled)
            self.settings_obj.save_settings()
            
    def show(self):
        self.refresh()
        self.update()
        super(Log_Settings_Window, self).show()
            
class TableModel(QtCore.QAbstractTableModel):

    def __init__(self, data, merged_cells):
        super().__init__()
        self._data = data if data else []
        self.merged_cells = merged_cells
        # Find the maximum column count across all rows to handle ragged data
        self._max_cols = 0
        if self._data:
            for row in self._data:
                if row:
                    self._max_cols = max(self._max_cols, len(row) if row else 0)
        print(f"DEBUG TableModel: data rows={len(self._data)}, max_cols={self._max_cols}")

    def data(self, index, role):
        if not index.isValid():
            return None
        
        if role == QtCore.Qt.ItemDataRole.DisplayRole:
            # Handle out-of-bounds or None values gracefully
            if index.row() >= len(self._data):
                return None
            row_data = self._data[index.row()]
            if row_data is None or index.column() >= len(row_data):
                return None
            
            value = row_data[index.column()]
            return value if value is not None else ""

    def rowCount(self, index):
        # The length of the outer list.
        return len(self._data) if self._data else 0

    def columnCount(self, index):
        # Return the maximum number of columns found across all rows
        return self._max_cols
        
if __name__ == '__main__':
    app = QApplication([])
    main = Log_Settings_Window()
    main.show()
    app.exec()